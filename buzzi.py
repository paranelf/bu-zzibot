import discord
import os
import openpyxl
from discord.ext import commands
import datetime

client = commands.Bot(command_prefix='!')

if os.path.exists("warnings"):
    print("Warnings Dir found, passing")
else:
    os.mkdir("warnings")


def find(filename, directory):
    if os.path.exists(directory + filename):
        return True
    else:
        return False


def returnAddData(filename, directory, num):
    try:
        f = open(directory + filename, "r")
        data = f.read()
        f.close()
        f = open(directory + filename, "w")
        f.write(str(int(data) + int(num)))
        f.close()
        return int(data) + int(num)
    except FileNotFoundError:
        print("Error : File not found")


@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game("1.1")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.command(name="청소", pass_context=True)
async def _clear(ctx, *, amount=5):
    await ctx.channel.purge(limit=amount)
    await ctx.send("메시지 삭제 완료하였습니다.")


@client.command(name="영구밴", pass_context=True)
async def _ban(ctx, *, user_name: discord.Member):
    await user_name.ban()
    await ctx.send(str(user_name) + "님이 영구강퇴되었습니다.")


@client.command(name="밴", pass_context=True)
async def _kick(ctx, *, user_name: discord.Member, reason=None):
    await user_name.kick()
    await ctx.send(str(user_name) + "님이 강퇴되었습니다.")


@client.command(name="경고", pass_context=True)
@commands.has_permissions(administrator=True)
async def _warn(ctx, counts, user_name: discord.Member = None, reason="없음"):
    if user_name == None or user_name == ctx.message.author:
        await ctx.send("자신에게 경고를 줄 수 없습니다.")
    else:
        foundfile = find(str(user_name) + ".txt", "warnings/")
        if foundfile:
            warnings = returnAddData(str(user_name) + ".txt", "warnings/", counts)
            if warnings >= 3:
                await user_name.ban()
                await ctx.send(str(user_name) + "이(가) 경고 누적 3회로 밴 되었습니다.")
            else:
                await ctx.send(str(user_name) + "님 경고입니다.")
        else:
            f = open("warnings/" + str(user_name) + ".txt", "w+")
            f.write(str(int(counts)))
            f.close()
            await ctx.send(str(user_name) + "님 경고입니다.")


async def _warn_error(ctx, error):
    if isinstance(error, commands.MissingPermissions):
        await ctx.send("{}님, 당신은 이 명령을 실행하실 권한이 없습니다.".format(ctx.message.author))

@client.event
async def on_member_join(member):
    channel = client.get_channel(701334526105878551)
    membername = discord.utils.get(member.guild.members, name=member.name)
    role = discord.utils.get(member.guild.roles, name="해머단")
    text = "{}님 사부찌의 디코서버에 오신것을 진심으로 환영한다고!!".format(membername.mention)
    date = datetime.datetime.utcfromtimestamp(((int(member.id) >> 22) + 1420070400000) / 1000)
    embed = discord.Embed(color=0xfbff00)
    embed.set_author(name=member.display_name, icon_url=member.avatar_url)
    embed.add_field(name="이름", value=member.name, inline=True)
    embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일",inline=True)
    embed.set_thumbnail(url=member.avatar_url)
    await member.add_roles(role)
    await channel.send(text, embed=embed)


@client.event
async def on_message(message):
    await client.process_commands(message)
    if message.content.startswith(message.content[0:]):
        file = openpyxl.load_workbook("추가.xlsx")
        sheet = file.active
        memory = message.content[0:]
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith("!추가"):
        file = openpyxl.load_workbook("추가.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("명령어가 정상적으로 추가되었습니다")
                break
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("명령어가 이미 존재합니다, 수정하실거면 !수정 을 삭제하실거면 !삭제 를 이용해주세요.")
                break
        file.save("추가.xlsx")

    if message.content.startswith("!수정"):
        file = openpyxl.load_workbook("추가.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("명령어가 정상적으로 수정 되었습니다.")
                break
            if sheet["A" + str(i)].value == "-":
                await message.channel.send("존재하지 않는 명령어 입니다.")
                break
        file.save("추가.xlsx")

    if message.content.startswith("!삭제"):
        file = openpyxl.load_workbook("추가.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                await message.channel.send("명령어가 정상적으로 삭제 되었습니다.")
                break
            if sheet["A" + str(i)].value == "-":
                await message.channel.send("존재하지 않는 명령어 입니다.")
                break
        file.save("추가.xlsx")

    if message.content.startswith("사하"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 트하 트하~~".format(membername.mention))
    if message.content.startswith("찌하"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 안뇽 안뇽~~ ".format(membername.mention))
    if message.content.startswith("사밤"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 사밤~~ 잘자용~".format(membername.mention))
    if message.content.startswith("사나잇"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 사나잇~~".format(membername.mention))
    if message.content.startswith("찌나잇"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 편안한 밤 되시길 바래용~~".format(membername.mention))
    if message.content.startswith("찌밤"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 내꿈꿔야됫!!!".format(membername.mention))
    if message.content.startswith("쫀아"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 좋은 아침 입니다 ㅎㅎ~~".format(membername.mention))
    if message.content.startswith("사모닝"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 잘잤어용?? ".format(membername.mention))
    if message.content.startswith("찌모닝"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 내꿈 꿨죠?".format(membername.mention))
    if message.content.startswith("사바"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 잘가요... ㅠㅠ".format(membername.mention))
    if message.content.startswith("찌바"):
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        await message.channel.send("{}님 조심히 들어가시고 다음에 봐용~~ 트바~~".format(membername.mention))
    if message.content.startswith("!내정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=0xfbff00)
        embed.set_author(name=message.author.display_name, icon_url=message.author.avatar_url)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일",
                        inline=True)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)
    if message.content.startswith("!명령어"):
        embed = discord.Embed(title="부찌봇명령어", color=0x0400ff)
        embed.add_field(name="관리자전용", value="!밴 @닉네임, !영구밴 @닉네임, !청소<지울 메시지 갯수>, !경고 @닉네임", inline=False)
        embed.add_field(name="채팅전용", value="사하, 찌하, 사바, 찌바, 사밤, 찌밤, 사나잇, 찌나잇, 쫀아, 사모닝, 찌모닝", inline=False)
        embed.add_field(name="명령어전용", value="!금지어, !추가, !수정, !삭제, !출첵", inline=False)
        await message.channel.send(embed=embed)

    if message.content.startswith("!출첵"):
        file = openpyxl.load_workbook("출석부.xlsx")
        sheet = file.active
        date = datetime.datetime.utcfromtimestamp(((int(message.id) >> 22) + 1420070400000) / 1000)
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        if message.channel.id != 748217990431899729:
            await message.channel.send("출첵은 출첵방에서만 사용해주세요!!")
        else:
            i = 1
            while True:
                if sheet["A" + str(i)].value == str(message.author.id):
                    if date == date:
                        await message.channel.send("{}님 오늘은 이미 출석체크를 하셨습니다, 내일 다시 시도해주세요.".format(membername.mention))
                    else:
                        sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                        file.save("출석부.xlsx")
                        await message.channel.send(
                            "{}님 ".format(membername.mention) + str(date.year) + "년" + str(date.month) + "월" + str(
                                date.day) + "일 출석체크가 완료되었습니다.(" + sheet["B" + str(i)].value + "회)")
                    break

                if sheet["A" + str(i)].value == None:
                    sheet["A" + str(i)].value = str(message.author.id)
                    sheet["B" + str(i)].value = 1
                    file.save("출석부.xlsx")
                    date = datetime.datetime.utcfromtimestamp(((int(message.id) >> 22) + 1420070400000) / 1000)
                    membername = discord.utils.get(message.guild.members, name=message.author.name)
                    await message.channel.send(
                        "{}님 ".format(membername.mention) + str(date.year) + "년" + str(date.month) + "월" + str(
                            date.day) + "일 출석체크가 완료되었습니다.(출첵횟수: " +  str(sheet["B" + str(i)].value) + "회)")
                    break
                i += 1    

    if message.content.startswith(message.content[0:]):
        file1 = openpyxl.load_workbook("금지어.xlsx")
        file2 = openpyxl.load_workbook("경고.xlsx")
        sheet1 = file1.active
        sheet2 = file2.active
        memory = message.content[0:]
        author1 = message.guild.get_member(int(message.author.id))
        membername = discord.utils.get(message.guild.members, name=message.author.name)
        for i in range(1, 51):
            if sheet1["A" + str(i)].value == memory:
                i = 1
                while True:
                    if sheet2["A" + str(i)].value == str(message.author.id):
                        sheet2["B" + str(i)].value = int(sheet2["B" + str(i)].value) + 1
                        file2.save("경고.xlsx")
                        if sheet2["B" + str(i)].value == 3:
                            await message.guild.kick(author1)
                            sheet2["B" + str(i)].value = int(sheet2["B" + str(i)].value) - 3
                            file2.save("경고.xlsx")
                            await message.channel.send(
                                "{}님은 금지어 사용으로 인해 경고가 3회 누적되어 본 서버에서 추방됩니다.".format(membername.mention))
                            await author1.send(
                                "{}님은 금지어 사용으로 인해 경고가 3회 누적되어 본 서버에서 추방되었습니다.".format(membername.mention))

                        else:
                            await message.channel.send(
                                "{}님 금지어 사용으로 경고 1회 입니다. ".format(membername.mention) + "(현재누적: " + str(
                                    sheet2["B" + str(i)].value) + "회)")
                        break
                    if sheet2["A" + str(i)].value == None:
                        sheet2["A" + str(i)].value = str(message.author.id)
                        sheet2["B" + str(i)].value = 1
                        file2.save("경고.xlsx")
                        await message.channel.send(
                            "{}님 금지어 사용으로 경고 1회 입니다. ".format(membername.mention) + "(현재누적: " + str(
                                sheet2["B" + str(i)].value) + "회)")
                        break
                    i += 1

     if message.content.startswith("!금지어 추가"):
        file = openpyxl.load_workbook("금지어.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                await message.channel.send("금지어목록에 금지어가 정상적으로 추가되었습니다.")
                break
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("금지어가 이미 존재합니다, 수정하실거면 !금지어 수정 을 삭제하실거면 !금지어 삭제 를 이용해주세요.")
                break  
        file.save("금지어.xlsx")
    
    if message.content.startswith("!금지어 수정"):
        file = openpyxl.load_workbook("금지어.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("명령어가 정상적으로 수정 되었습니다.")
                break
            if sheet["A" + str(i)].value == "-":
                await message.channel.send("존재하지 않는 금지어 입니다.")
                break
        file.save("금지어.xlsx")

    if message.content.startswith("!금지어 삭제"):
        file = openpyxl.load_workbook("금지어.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                await message.channel.send("금지어가 정상적으로 삭제 되었습니다.")
                break
            if sheet["A" + str(i)].value == "-":
                await message.channel.send("존재하지 않는 금지어 입니다.")
                break
        file.save("금지어.xlsx")
    
   


access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
